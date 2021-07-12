from flask import Flask, request, jsonify

app = Flask(__name__)


@app.route('/')
def cc():
    return 'HAI'


@app.route("/show_json", methods=['GET', 'POST'])
def show_json(self=''):
    from_postman = request.get_json()
    a = {}
    for i, j in dict(from_postman).items():
        if isinstance(j, dict):
            for k, l in j.items():
                if k == "kamal":
                    a[k] = "Ulahanayagan"
                if k == "Rajini":
                    a[k] = "Superstarr"
        elif isinstance(j, list):
            for k in j:
                if k == "Kamal":
                    a[str(k)] = "Ulahanayagan"
                if k == "Rajini":
                    a[str(k)] = "Superstarr"
    from_postman['det'] = a
    return jsonify(from_postman)


@app.route("/upload_excel", methods=['GET', 'POST'])
def excel_to_mysql(col=1, start_row=2, end_row=0):
    fil_from_post_man = request.files['file']
    filename=fil_from_post_man.filename
    ext=str(filename).split('.')
    if ext[1]=='tsv':
        import csv
        fil_from_post_man.save("D:\\sriram\\"+filename)
        file = open("D:\\sriram\\"+filename)
        file1 = csv.reader(file, delimiter="\t")
        header = [j for i, j in enumerate(file1)]
        realheader = header[0]
        val = []
        for i in header:
            val.append(dict(zip(realheader, i)))
        return jsonify(val[1:])
    else:

        from openpyxl import load_workbook
        ws = load_workbook(fil_from_post_man)
        respond_value={}
        value=[]
        show = ws.active
        for i in show.iter_rows(max_row=col, values_only=True):
            heading=list(i)
        for i in show.iter_rows(min_row=start_row, max_row=show.max_row if end_row > 0 else end_row, values_only=True):
            value.append(dict(zip(heading,list(i))))
        respond_value["file"]=value
        return jsonify(respond_value)


@app.route("/json_return",methods=["get"])
def json_return():
    from_postman = request.get_json()
    lst = []
    for i in from_postman["json"]:
        gstamount = (i["gross"] / 100) * i["gst%"]
        i["gstamount"] = gstamount
        i["cgst"] = gstamount / 2
        i["sgst"] = i["cgst"]
        lst.append(i)
    j = {"json-return": lst}
    from cryptography.fernet import Fernet
    message =str(j).encode()
    fernet_class = Fernet("2U_YYVtUzTHUkjictbju5Uy2Z6eDbLBpAAHHTKOrvJo=")
    encrypted = fernet_class.encrypt(message)
    with open("D:\sriram\sri_script\sri_encript\sri.json", "wb") as key_file:
        key_file.write(encrypted)
    return jsonify(j)




if __name__ == '__main__':
    app.debug = True
    app.run()
